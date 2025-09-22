from celery import shared_task
from django.utils import timezone
from django.conf import settings
import openai
import anthropic
import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
from .models import AnalysisTask, AnalysisResult, AnalysisExport
from apps.github.models import Commit
from utils.encryption import decrypt_data


class AIAnalysisClient:
    """Client for AI analysis using OpenAI or Anthropic"""
    
    def __init__(self, provider, model, api_key, temperature=0.7, max_tokens=2000):
        self.provider = provider
        self.model = model
        self.temperature = temperature
        self.max_tokens = max_tokens
        
        if provider == 'openai':
            openai.api_key = api_key
        elif provider == 'anthropic':
            self.anthropic_client = anthropic.Anthropic(api_key=api_key)
    
    def analyze(self, system_prompt, analysis_prompt):
        """Run AI analysis"""
        try:
            if self.provider == 'openai':
                response = openai.ChatCompletion.create(
                    model=self.model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": analysis_prompt}
                    ],
                    temperature=self.temperature,
                    max_tokens=self.max_tokens
                )
                
                return {
                    'analysis': response.choices[0].message.content,
                    'tokens_used': response.usage.total_tokens,
                    'cost': self._calculate_openai_cost(response.usage.total_tokens)
                }
            
            elif self.provider == 'anthropic':
                response = self.anthropic_client.messages.create(
                    model=self.model,
                    max_tokens=self.max_tokens,
                    temperature=self.temperature,
                    system=system_prompt,
                    messages=[
                        {"role": "user", "content": analysis_prompt}
                    ]
                )
                
                return {
                    'analysis': response.content[0].text,
                    'tokens_used': response.usage.input_tokens + response.usage.output_tokens,
                    'cost': self._calculate_anthropic_cost(response.usage.input_tokens, response.usage.output_tokens)
                }
            
        except Exception as e:
            raise Exception(f"AI analysis failed: {str(e)}")
    
    def _calculate_openai_cost(self, tokens):
        """Calculate OpenAI API cost (approximate)"""
        # These are example rates - update with actual pricing
        rates = {
            'gpt-3.5-turbo': 0.002 / 1000,  # $0.002 per 1K tokens
            'gpt-4': 0.03 / 1000,           # $0.03 per 1K tokens
            'gpt-4-turbo': 0.01 / 1000,     # $0.01 per 1K tokens
        }
        return tokens * rates.get(self.model, 0.002 / 1000)
    
    def _calculate_anthropic_cost(self, input_tokens, output_tokens):
        """Calculate Anthropic API cost (approximate)"""
        # These are example rates - update with actual pricing
        rates = {
            'claude-3-haiku': {'input': 0.00025 / 1000, 'output': 0.00125 / 1000},
            'claude-3-sonnet': {'input': 0.003 / 1000, 'output': 0.015 / 1000},
            'claude-3-opus': {'input': 0.015 / 1000, 'output': 0.075 / 1000},
        }
        
        model_rates = rates.get(self.model, rates['claude-3-sonnet'])
        return (input_tokens * model_rates['input']) + (output_tokens * model_rates['output'])


@shared_task(bind=True)
def run_analysis_task(self, task_id):
    """Run AI analysis task"""
    try:
        task = AnalysisTask.objects.get(id=task_id)
        task.status = 'processing'
        task.started_at = timezone.now()
        task.save(update_fields=['status', 'started_at'])
        
        # Update progress
        task.update_progress(10, "Fetching commit data...")
        
        # Get commits for analysis
        commits_query = Commit.objects.filter(repository=task.repository)
        
        if task.date_from:
            commits_query = commits_query.filter(committed_at__gte=task.date_from)
        if task.date_to:
            commits_query = commits_query.filter(committed_at__lte=task.date_to)
        if task.developer_filter:
            commits_query = commits_query.filter(author_name__icontains=task.developer_filter)
        
        commits = commits_query.order_by('-committed_at')
        
        if not commits.exists():
            raise Exception("No commits found for the specified criteria")
        
        task.update_progress(30, "Processing commit data...")
        
        # Prepare commit data for analysis
        commit_data = []
        total_additions = 0
        total_deletions = 0
        total_files_changed = 0
        developer_stats = {}
        
        for commit in commits:
            commit_info = {
                'sha': commit.sha,
                'message': commit.message,
                'author': commit.author_name,
                'date': commit.committed_at.isoformat(),
                'additions': commit.additions,
                'deletions': commit.deletions,
                'files_changed': commit.files_changed
            }
            commit_data.append(commit_info)
            
            # Update totals
            total_additions += commit.additions
            total_deletions += commit.deletions
            total_files_changed += commit.files_changed
            
            # Update developer stats
            author = commit.author_name
            if author not in developer_stats:
                developer_stats[author] = {
                    'commits': 0,
                    'additions': 0,
                    'deletions': 0,
                    'files_changed': 0
                }
            
            developer_stats[author]['commits'] += 1
            developer_stats[author]['additions'] += commit.additions
            developer_stats[author]['deletions'] += commit.deletions
            developer_stats[author]['files_changed'] += commit.files_changed
        
        task.update_progress(50, "Preparing AI analysis...")
        
        # Get user's API key
        user = task.user
        if task.agent.ai_provider == 'openai':
            api_key = decrypt_data(user.openai_api_key) if user.openai_api_key else None
        else:
            api_key = decrypt_data(user.anthropic_api_key) if user.anthropic_api_key else None
        
        if not api_key:
            raise Exception(f"No API key found for {task.agent.ai_provider}")
        
        # Initialize AI client
        ai_client = AIAnalysisClient(
            provider=task.agent.ai_provider,
            model=task.agent.model_name,
            api_key=api_key,
            temperature=task.agent.temperature,
            max_tokens=task.agent.max_tokens
        )
        
        task.update_progress(60, "Running AI analysis...")
        
        # Prepare analysis data
        analysis_data = {
            'repository_name': task.repository.name,
            'developer_name': task.developer_filter or 'All developers',
            'date_range': f"{task.date_from.strftime('%Y-%m-%d') if task.date_from else 'Beginning'} to {task.date_to.strftime('%Y-%m-%d') if task.date_to else 'Now'}",
            'total_commits': len(commit_data),
            'total_additions': total_additions,
            'total_deletions': total_deletions,
            'file_changes': total_files_changed,
            'commit_data': json.dumps(commit_data[:50], indent=2)  # Limit to 50 commits for prompt
        }
        
        # Format prompts
        system_prompt = task.agent.format_system_prompt(analysis_data)
        analysis_prompt = task.agent.format_prompt(analysis_data)
        
        # Run AI analysis
        ai_response = ai_client.analyze(system_prompt, analysis_prompt)
        
        task.update_progress(80, "Saving analysis results...")
        
        # Create analysis result
        result = AnalysisResult.objects.create(
            task=task,
            raw_analysis=ai_response['analysis'],
            formatted_analysis=ai_response['analysis'],  # Could be formatted differently
            total_commits=len(commit_data),
            total_additions=total_additions,
            total_deletions=total_deletions,
            total_files_changed=total_files_changed,
            developer_stats=developer_stats,
            tokens_used=ai_response['tokens_used'],
            analysis_cost=ai_response['cost']
        )
        
        task.update_progress(100, "Analysis completed successfully")
        task.status = 'completed'
        task.completed_at = timezone.now()
        task.save(update_fields=['status', 'completed_at'])
        
        return f"Analysis completed successfully for task {task_id}"
        
    except Exception as e:
        task = AnalysisTask.objects.get(id=task_id)
        task.status = 'failed'
        task.error_message = str(e)
        task.completed_at = timezone.now()
        task.save(update_fields=['status', 'error_message', 'completed_at'])
        
        raise Exception(f"Analysis task failed: {str(e)}")


@shared_task(bind=True)
def create_export_task(self, export_id):
    """Create analysis export file"""
    try:
        export = AnalysisExport.objects.get(id=export_id)
        export.status = 'processing'
        export.save(update_fields=['status'])
        
        result = export.analysis_result
        task = result.task
        
        # Create export directory if it doesn't exist
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        filename = f"analysis_export_{export.id}.{export.format}"
        file_path = os.path.join(export_dir, filename)
        
        if export.format == 'excel':
            create_excel_export(export, result, file_path)
        elif export.format == 'csv':
            create_csv_export(export, result, file_path)
        elif export.format == 'json':
            create_json_export(export, result, file_path)
        elif export.format == 'pdf':
            create_pdf_export(export, result, file_path)
        
        # Update export record
        file_size = os.path.getsize(file_path)
        export.status = 'completed'
        export.file_path = filename
        export.file_size = file_size
        export.completed_at = timezone.now()
        export.expires_at = timezone.now() + timedelta(days=7)  # Expire after 7 days
        export.save(update_fields=['status', 'file_path', 'file_size', 'completed_at', 'expires_at'])
        
        return f"Export created successfully: {filename}"
        
    except Exception as e:
        export = AnalysisExport.objects.get(id=export_id)
        export.status = 'failed'
        export.save(update_fields=['status'])
        
        raise Exception(f"Export creation failed: {str(e)}")


def create_excel_export(export, result, file_path):
    """Create Excel export with charts and formatting"""
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Analysis Summary"
    
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add summary data
    ws_summary['A1'] = "GitHub Repository Analysis Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    
    ws_summary['A3'] = "Repository:"
    ws_summary['B3'] = result.task.repository.name
    ws_summary['A4'] = "Analysis Date:"
    ws_summary['B4'] = result.created_at.strftime('%Y-%m-%d %H:%M')
    ws_summary['A5'] = "Agent Used:"
    ws_summary['B5'] = result.task.agent.name
    
    ws_summary['A7'] = "Statistics"
    ws_summary['A7'].font = header_font
    
    ws_summary['A8'] = "Total Commits:"
    ws_summary['B8'] = result.total_commits
    ws_summary['A9'] = "Total Additions:"
    ws_summary['B9'] = result.total_additions
    ws_summary['A10'] = "Total Deletions:"
    ws_summary['B10'] = result.total_deletions
    ws_summary['A11'] = "Files Changed:"
    ws_summary['B11'] = result.total_files_changed
    
    # Analysis sheet
    ws_analysis = wb.create_sheet("AI Analysis")
    ws_analysis['A1'] = "AI Analysis Result"
    ws_analysis['A1'].font = header_font
    ws_analysis['A3'] = result.formatted_analysis
    ws_analysis['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Developer stats sheet
    if result.developer_stats:
        ws_devs = wb.create_sheet("Developer Statistics")
        ws_devs['A1'] = "Developer Statistics"
        ws_devs['A1'].font = header_font
        
        # Headers
        headers = ['Developer', 'Commits', 'Additions', 'Deletions', 'Files Changed']
        for col, header in enumerate(headers, 1):
            cell = ws_devs.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data
        for row, (dev, stats) in enumerate(result.developer_stats.items(), 4):
            ws_devs.cell(row=row, column=1, value=dev)
            ws_devs.cell(row=row, column=2, value=stats['commits'])
            ws_devs.cell(row=row, column=3, value=stats['additions'])
            ws_devs.cell(row=row, column=4, value=stats['deletions'])
            ws_devs.cell(row=row, column=5, value=stats['files_changed'])
        
        # Add chart if requested
        if export.include_charts:
            chart = BarChart()
            chart.title = "Commits by Developer"
            chart.x_axis.title = "Developer"
            chart.y_axis.title = "Commits"
            
            data = Reference(ws_devs, min_col=2, min_row=3, max_row=3+len(result.developer_stats), max_col=2)
            cats = Reference(ws_devs, min_col=1, min_row=4, max_row=3+len(result.developer_stats))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_devs.add_chart(chart, "G3")
    
    # Raw data sheet (if requested)
    if export.include_raw_data:
        ws_raw = wb.create_sheet("Raw Data")
        ws_raw['A1'] = "Raw Analysis Data"
        ws_raw['A1'].font = header_font
        ws_raw['A3'] = result.raw_analysis
        ws_raw['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(file_path)


def create_csv_export(export, result, file_path):
    """Create CSV export"""
    data = []
    
    # Summary data
    data.append(['Repository', result.task.repository.name])
    data.append(['Analysis Date', result.created_at.strftime('%Y-%m-%d %H:%M')])
    data.append(['Agent Used', result.task.agent.name])
    data.append(['Total Commits', result.total_commits])
    data.append(['Total Additions', result.total_additions])
    data.append(['Total Deletions', result.total_deletions])
    data.append(['Files Changed', result.total_files_changed])
    data.append([])  # Empty row
    
    # Developer stats
    if result.developer_stats:
        data.append(['Developer Statistics'])
        data.append(['Developer', 'Commits', 'Additions', 'Deletions', 'Files Changed'])
        
        for dev, stats in result.developer_stats.items():
            data.append([dev, stats['commits'], stats['additions'], stats['deletions'], stats['files_changed']])
        
        data.append([])  # Empty row
    
    # Analysis result
    data.append(['AI Analysis'])
    data.append([result.formatted_analysis])
    
    df = pd.DataFrame(data)
    df.to_csv(file_path, index=False, header=False)


def create_json_export(export, result, file_path):
    """Create JSON export"""
    data = {
        'repository': result.task.repository.name,
        'analysis_date': result.created_at.isoformat(),
        'agent_used': result.task.agent.name,
        'statistics': {
            'total_commits': result.total_commits,
            'total_additions': result.total_additions,
            'total_deletions': result.total_deletions,
            'files_changed': result.total_files_changed
        },
        'developer_stats': result.developer_stats,
        'ai_analysis': result.formatted_analysis,
        'tokens_used': result.tokens_used,
        'analysis_cost': float(result.analysis_cost)
    }
    
    if export.include_raw_data:
        data['raw_analysis'] = result.raw_analysis
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def create_pdf_export(export, result, file_path):
    """Create PDF export (placeholder - would need reportlab or similar)"""
    # This would require additional dependencies like reportlab
    # For now, create a simple text file
    with open(file_path.replace('.pdf', '.txt'), 'w', encoding='utf-8') as f:
        f.write(f"GitHub Repository Analysis Report\n")
        f.write(f"================================\n\n")
        f.write(f"Repository: {result.task.repository.name}\n")
        f.write(f"Analysis Date: {result.created_at.strftime('%Y-%m-%d %H:%M')}\n")
        f.write(f"Agent Used: {result.task.agent.name}\n\n")
        f.write(f"Statistics:\n")
        f.write(f"- Total Commits: {result.total_commits}\n")
        f.write(f"- Total Additions: {result.total_additions}\n")
        f.write(f"- Total Deletions: {result.total_deletions}\n")
        f.write(f"- Files Changed: {result.total_files_changed}\n\n")
        f.write(f"AI Analysis:\n")
        f.write(f"{result.formatted_analysis}\n")
    
    # Update file path to .txt since we're not creating actual PDF
    export.file_path = export.file_path.replace('.pdf', '.txt')